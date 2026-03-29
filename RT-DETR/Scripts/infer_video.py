import cv2
import torch
import time
import sys
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG — only edit this section ──────────────────────────────────────────
CONFIG     = r"configs\rtdetrv2\rtdetrv2_r50vd_drone.yml"
CHECKPOINT = r"configs\best.pth"
VIDEO_IN   = r"Inference\phantom14.mp4" 
VIDEO_OUT  = r"Inference\output14_infer.mp4"
EXCEL_OUT  = r"Inference\infer14_stats.xlsx"
THRESHOLD  = 0.4
INPUT_SIZE = 640    # model trained at 640×640
DEVICE     = "cuda"
# ─────────────────────────────────────────────────────────────────────────────

RTDETR_PATH = "/home/devananth/office/Vision/RT-DETR/rtdetrv2_pytorch"
sys.path.insert(0, RTDETR_PATH)

from src.core import YAMLConfig

# ── LOAD MODEL ────────────────────────────────────────────────────────────────
device = torch.device(DEVICE if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

cfg = YAMLConfig(CONFIG, resume=CHECKPOINT)
checkpoint = torch.load(CHECKPOINT, map_location=device)
state = checkpoint.get("ema", {}).get("module", checkpoint.get("model", checkpoint))
cfg.model.load_state_dict(state)

model = cfg.model.to(device)
model.eval()

postprocessor = cfg.postprocessor
postprocessor.deploy()

print("Model loaded successfully")

# ── VIDEO SETUP ───────────────────────────────────────────────────────────────
cap = cv2.VideoCapture(VIDEO_IN)
assert cap.isOpened(), f"Cannot open video: {VIDEO_IN}"

total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
fps          = cap.get(cv2.CAP_PROP_FPS)
W            = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
H            = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

out = cv2.VideoWriter(
    VIDEO_OUT,
    cv2.VideoWriter_fourcc(*"mp4v"),
    fps, (W, H)   # output at original resolution
)

print(f"Video  : {W}x{H} @ {fps:.1f}fps | {total_frames} frames")
print(f"Model input: {INPUT_SIZE}x{INPUT_SIZE}")

# ── INFERENCE LOOP ────────────────────────────────────────────────────────────
results   = []
frame_num = 0

with torch.no_grad():
    while True:
        ret, frame = cap.read()
        if not ret:
            break

        # Preprocess — resize to 640×640 for model input
        img_rgb     = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img_resized = cv2.resize(img_rgb, (INPUT_SIZE, INPUT_SIZE))
        img_tensor  = torch.from_numpy(img_resized).permute(2, 0, 1).float() / 255.0
        img_tensor  = img_tensor.unsqueeze(0).to(device)

        # Pass model input size (not original) as orig_size
        orig_size = torch.tensor([[INPUT_SIZE, INPUT_SIZE]], device=device)

        # Inference
        t0 = time.perf_counter()
        outputs = model(img_tensor)
        labels, boxes, scores = postprocessor(outputs, orig_size)
        torch.cuda.synchronize()
        t1 = time.perf_counter()

        inference_ms = (t1 - t0) * 1000

        # Filter by threshold
        mask        = scores[0] > THRESHOLD
        kept_scores = scores[0][mask].cpu().numpy()
        kept_boxes  = boxes[0][mask].cpu().numpy()   # boxes in 640×640 space

        n_objects = int(mask.sum())
        avg_conf  = float(kept_scores.mean()) if n_objects > 0 else 0.0
        max_conf  = float(kept_scores.max())  if n_objects > 0 else 0.0

        results.append({
            "frame":        frame_num,
            "inference_ms": round(inference_ms, 3),
            "n_objects":    n_objects,
            "avg_conf":     round(avg_conf, 4),
            "max_conf":     round(max_conf, 4),
        })

        # Scale boxes back to original frame resolution for drawing
        scale_x = W / INPUT_SIZE
        scale_y = H / INPUT_SIZE

        for i, box in enumerate(kept_boxes):
            x1 = int(box[0] * scale_x)
            y1 = int(box[1] * scale_y)
            x2 = int(box[2] * scale_x)
            y2 = int(box[3] * scale_y)
            conf = kept_scores[i]
            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.putText(frame, f"drone {conf:.2f}",
                        (x1, max(y1 - 8, 0)),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

        # HUD
        cv2.putText(frame,
                    f"Frame: {frame_num}  |  Objects: {n_objects}  |  {inference_ms:.1f} ms",
                    (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 200, 255), 2)

        out.write(frame)
        frame_num += 1

        if frame_num % 100 == 0:
            print(f"  Processed {frame_num}/{total_frames} frames")

cap.release()
out.release()
print(f"\nVideo saved → {VIDEO_OUT}")

# ── BUILD EXCEL ───────────────────────────────────────────────────────────────
wb = Workbook()

# Sheet 1: Per-frame data
ws = wb.active
ws.title = "Per Frame Stats"

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
ALT_FILL    = PatternFill("solid", start_color="D6E4F0")
NORMAL_FONT = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
thin        = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

headers    = ["Frame #", "Inference Time (ms)", "No. of Objects", "Avg Confidence", "Max Confidence"]
col_widths = [12, 22, 18, 18, 18]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell           = ws.cell(row=1, column=col, value=h)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = CENTER
    cell.border    = BORDER
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 20

for i, r in enumerate(results):
    row    = i + 2
    values = [r["frame"], r["inference_ms"], r["n_objects"], r["avg_conf"], r["max_conf"]]
    fill   = ALT_FILL if i % 2 == 0 else None
    for col, val in enumerate(values, 1):
        cell           = ws.cell(row=row, column=col, value=val)
        cell.font      = NORMAL_FONT
        cell.alignment = CENTER
        cell.border    = BORDER
        if fill:
            cell.fill = fill

# Sheet 2: Summary
ws2 = wb.create_sheet("Summary")

n         = len(results)
inf_times = [r["inference_ms"] for r in results]
confs_avg = [r["avg_conf"]     for r in results if r["n_objects"] > 0]
confs_max = [r["max_conf"]     for r in results if r["n_objects"] > 0]
n_objects = [r["n_objects"]    for r in results]

summary = [
    ("Video File",                       Path(VIDEO_IN).name),
    ("Total Frames Processed",           n),
    ("Original Resolution",              f"{W}x{H}"),
    ("Model Input Size",                 f"{INPUT_SIZE}x{INPUT_SIZE}"),
    ("Video FPS",                        round(fps, 2)),
    ("Confidence Threshold",             THRESHOLD),
    ("Device",                           str(device)),
    ("Checkpoint",                       Path(CHECKPOINT).name),
    ("", ""),
    ("── Inference Time ──",             ""),
    ("Avg Inference Time (ms)",          round(np.mean(inf_times), 3)),
    ("Min Inference Time (ms)",          round(np.min(inf_times),  3)),
    ("Max Inference Time (ms)",          round(np.max(inf_times),  3)),
    ("Std Inference Time (ms)",          round(np.std(inf_times),  3)),
    ("Effective FPS",                    round(1000 / np.mean(inf_times), 1)),
    ("", ""),
    ("── Detection Stats ──",            ""),
    ("Total Detections",                 sum(n_objects)),
    ("Avg Objects / Frame",              round(np.mean(n_objects), 3)),
    ("Max Objects in a Frame",           int(np.max(n_objects))),
    ("Frames with Detection",            sum(1 for x in n_objects if x > 0)),
    ("Frames without Detection",         sum(1 for x in n_objects if x == 0)),
    ("", ""),
    ("── Confidence ──",                 ""),
    ("Avg Confidence (detected frames)", round(np.mean(confs_avg), 4) if confs_avg else "N/A"),
    ("Max Confidence (any frame)",       round(np.max(confs_max), 4) if confs_max else "N/A"),
]

LABEL_FILL   = PatternFill("solid", start_color="1F4E79")
SECTION_FILL = PatternFill("solid", start_color="2E75B6")
VALUE_FILL   = PatternFill("solid", start_color="EBF3FB")

ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 25

for row, (label, value) in enumerate(summary, 1):
    lc = ws2.cell(row=row, column=1, value=label)
    vc = ws2.cell(row=row, column=2, value=value)
    lc.border = BORDER
    vc.border = BORDER

    if label.startswith("──"):
        lc.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        lc.fill = SECTION_FILL
        vc.fill = SECTION_FILL
    elif label == "":
        pass
    elif row <= 8:
        lc.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        lc.fill = LABEL_FILL
        vc.font = Font(name="Arial", size=10, bold=True)
        vc.fill = VALUE_FILL
        vc.alignment = CENTER
    else:
        lc.font = Font(name="Arial", size=10)
        vc.font = Font(name="Arial", size=10)
        vc.fill = VALUE_FILL
        vc.alignment = CENTER

    lc.alignment = Alignment(vertical="center")
    ws2.row_dimensions[row].height = 18

wb.save(EXCEL_OUT)
print(f"Excel saved  → {EXCEL_OUT}")
print(f"\n── Summary ──────────────────────────────")
print(f"  Frames processed : {n}")
print(f"  Avg inf time     : {np.mean(inf_times):.2f} ms")
print(f"  Effective FPS    : {1000/np.mean(inf_times):.1f}")
print(f"  Max inf time     : {np.max(inf_times):.2f} ms")
print(f"  Total detections : {sum(n_objects)}")
print(f"  Frames w/ drone  : {sum(1 for x in n_objects if x > 0)}")
print(f"  Frames w/o drone : {sum(1 for x in n_objects if x == 0)}")